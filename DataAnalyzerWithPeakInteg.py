##!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Last Modified: April 7, 2021
@author: sarahzeichner

This code has all of the data processing code, to take data after it has been processed by FT statistic 
(or equivalent) and calculate isotope ratios based on input
"""

import matplotlib
import csv
import os
import numpy as np
import pandas as pd
import math 
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import io
from collections import Counter
from scipy.stats import norm
from scipy.optimize import curve_fit
import scipy.integrate 
import openpyxl

#####################################################################
########################## CONSTANTS ################################
#####################################################################

WINDOW_LENGTH  = 5
SLOPE_THRESHHOLD = 0.008
NAN_REPLACER = 0.0000001
TRAP_RULE_BOOL = False

#####################################################################
########################## FUNCTIONS ################################
#####################################################################

def import_Peaks_From_FTStatFile(inputFileName):
    '''
    Import peaks from FT statistic output file into a workable form, step 1
    
    Inputs:
        inputFileName: The excel file to input from
        
    Outputs:
        A list, containing dictionaries for each mass with a set of peaks in the excel file. 
        The dictionaries have entries for 'tolerance', 'lastScan', 'refMass', and 'scans'. The 'scans' key directs to another list; 
        this has a dictionary for each indvidual scan, giving a bunch of data about that scan. 
    '''

    wb = openpyxl.load_workbook(inputFileName)
    ws = wb.active
    
    # list containing packets containing dicts of microscans for each measured peak
    peaks = []
    onMicroScans = False    

    for i, row in enumerate(ws.iter_rows(min_row=1,max_row=ws.max_row, values_only=True)):
        r = i+1 # r is row number in spreadsheet
        if 'Tolerance:' in row:
            peaks.append({})
            # Get the tolerance
            try:
                tol = float(ws.cell(r,2).value.strip(' ppm'))
            except:
                tol = float(ws.cell(r,2).value.strip(' mmu'))
            # Get the last scan of microscan packets
            lastScan = int(ws.cell(r,8).value)
            # Get the ref mass
            refMass = float(ws.cell(r,10).value)
            peaks[-1] = {'tolerance': tol, 'lastScan': lastScan,
                      'refMass': refMass, 'scans': []}
            continue
            
        if 'Measured Mass:' in row:
            # Saving rows to know what goes in each column (in case of changes later in the sheet)
            colIndex = row
            onMicroScans = True
            continue
                    
        if onMicroScans:
            if 'Aver:' in row:
                onMicroScans = False
                continue
            measuredMass    = ws.cell(r, colIndex.index('Measured Mass:')+1).value
            retTime         = ws.cell(r, colIndex.index('Ret. Time:')+1).value
            scanNumber      = ws.cell(r, colIndex.index('Scan Number:')+1).value
            absIntensity    = ws.cell(r, colIndex.index('Abs. Intensity:')+1).value
            integrationTime = ws.cell(r, colIndex.index('IT [ms]:')+1).value
            ftResolution    = ws.cell(r, colIndex.index('FT Resolution:')+1).value
            peakNoise       = ws.cell(r, colIndex.index('Peak Noise')+1).value
            totalIonCount   = ws.cell(r, colIndex.index('TIC:')+1).value
            ticTimesIT      = ws.cell(r, colIndex.index('TIC*IT:')+1).value
            peakResolution  = ws.cell(r, colIndex.index('Peak Resolution')+1).value
            peakBaseline    = ws.cell(r, colIndex.index('Peak Baseline')+1).value
            if measuredMass == '' and retTime == '' and scanNumber == '':
                continue
            if scanNumber == lastScan:
                onMicroScans = False
            peaks[-1]['scans'].append(({'mass': measuredMass, 'retTime': retTime, 'tic': totalIonCount,
                                        'scanNumber': scanNumber, 'absIntensity': absIntensity, 'integTime': integrationTime,'TIC*IT': ticTimesIT,'ftRes': ftResolution, 'peakNoise': peakNoise, 'peakRes': peakResolution, 'peakBase': peakBaseline}))
    return(peaks)

def convert_To_Pandas_DataFrame(peaks):
    '''
    Import peaks from FT statistic output file into a workable form, step 2
    
    Inputs:
        peaks: The peaks output from _importPeaksFromFTStatistic; a list of dictionaries. 
        
    Outputs: 
        A list, where each element is a pandas dataframe for an individual peak extracted by FTStatistic (i.e. a single line in the FTStat input .txt file). 
    '''
    rtnAllPeakDF = []

    for peak in peaks:
        try:
            columnLabels = list(peak['scans'][0])
            data = np.zeros((len(peak['scans']), len(columnLabels)))
        except:
            print("Could not find peak " + str(peak))
            continue
        # putting all scan data into an array
        for i in range(len(peak['scans'])):
            for j in range(len(columnLabels)):
                data[i, j] = peak['scans'][i][columnLabels[j]]
        # scan numbers as separate array for indices
        scanNumbers = data[:, columnLabels.index('scanNumber')]
        # constructing data frame
        peakDF = pd.DataFrame(data, index=scanNumbers, columns=columnLabels)

        # add it to the return pandas DF
        rtnAllPeakDF.append(peakDF)

    return(rtnAllPeakDF)

def calculate_Counts_And_ShotNoise(peakDF,resolution=120000,CN=4.4,z=1,Microscans=1):
    '''
    Calculate counts of each scan peak
    
    Inputs: 
        peakDf: An individual dataframe consisting of a single peak extracted by FTStatistic.
        CN: A factor from the 2017 paper to convert intensities into counts
        resolution: A reference resolution, for the same purpose (do not change!)
        z: The charge of the ion, used to convert intensities into counts
        Microscans: The number of scans a cycle is divided into, typically 1.
        
    Outputs: 
        The inputDF, with a column for 'counts' added. 
    '''
    #NOTE: Uncomment this to test just NL score for ratios
    #peakDF['counts'] = peakDF['absIntensity']  

    peakDF['counts'] = (peakDF['absIntensity'] /
                  peakDF['peakNoise']) * (CN/z) *(resolution/peakDF['ftRes'])**(0.5) * Microscans**(0.5)
    return peakDF

def calc_Append_Ratios(singleDf, allBelowOne = True, isotopeList = ['UnSub', '15N',  '13C']):
    '''
    Calculates both 15N and 13C ratios, writes them such that they are < 1, and adds them to the dataframe.
    Inputs:                               
            singleDF: An individual pandas dataframe, consisting of multiple peaks from FTStat combined into one dataframe by the _combinedSubstituted function.
            allBelowOne: if True, outputs ratios as 'Sub/unSub' or 'unSub/Sub', whichever is below 1. If false, outputs
            all as 'sub/unSub'. 
            isotopeList: A list of isotopes corresponding to the peaks extracted by FTStat, in the order they were extracted. 
                            This must be the same for each fragment. This is used to determine all ratios of interest, i.e. 13C/UnSub, and label them in the proper order. 
            
    Outputs:
            The dataframe with ratios added. It computes all ratios, because why not. 
    '''

    if allBelowOne:
        for i in range(len(isotopeList)):
            for j in range(len(isotopeList)):
                if j>i:
                    #determine which has more counts, set ratio accordingly
                    if singleDf['counts' + isotopeList[i]].sum() <= singleDf['counts' + isotopeList[j]].sum():
                        singleDf[isotopeList[i] + '/' + isotopeList[j]] = singleDf['counts' + isotopeList[i]] / singleDf['counts' + isotopeList[j]]
                        
                        #output timing of maximum peak to console for debugging
                        print(str(isotopeList[i]) + "timing: " + str(singleDf[['counts' + isotopeList[i]]].idxmax()))
                        print(str(isotopeList[j]) + "timing: " + str(singleDf[['counts' + isotopeList[j]]].idxmax()))

                    else:
                        singleDf[isotopeList[j] + '/' + isotopeList[i]] = singleDf['counts' + isotopeList[j]] / singleDf['counts' + isotopeList[i]]

                        #output timing of maximum peak to console for debugging
                        print(str(isotopeList[i]) + "timing: " + str(singleDf[['counts' + isotopeList[i]]].idxmax()))
                        print(str(isotopeList[j]) + "timing: " + str(singleDf[['counts' + isotopeList[j]]].idxmax()))
    else:
        for i in range(len(isotopeList)):
            for j in range(len(isotopeList)):
                if j>i:
                    singleDf[isotopeList[i] + '/' + isotopeList[j]] = singleDf['counts' + isotopeList[i]] / singleDf['counts' + isotopeList[j]]

                    #output timing of maximum peak to console for debugging
                    print(str(isotopeList[i]) + "timing: " + str(singleDf[['counts' + isotopeList[i]]].idxmax()))
                    print(str(isotopeList[j]) + "timing: " + str(singleDf[['counts' + isotopeList[j]]].idxmax()))
    return singleDf

def combine_Substituted_Peaks(peakDF, cullOn = [], cullZeroScansOn = False, baselineCorrectionOn=False, \
                            gc_elution_on = False, gc_elution_times = [], cullAmount = 2, isotopeList = ['13C','15N','UnSub'], \
                            minNL_over_maxNL = 0):
    '''
    Merge all extracted peaks from a given fragment into a single dataframe. For example, if I extracted six peaks, the 13C, 15N, and unsubstituted of fragments at 119 and 109, 
    this would input a list of six dataframes (one per peak) and combine them into two dataframes (one per fragment), each including data from the 13C, 15N, 
    and unsubstituted peaks of that fragment.
    
    Inputs: 
        peakDF: A list of dataframes. The list is the output of the _convertToPandasDataFrame function, and containts
        an individual dataframe for each peak extracted with FTStatistic. 
        cullOn: A target variable, like 'tic', or 'TIC*IT' to use to determine which scans to call. 
        cullZeroScansOn: Toggle whether or not you want to cull out zero scan counts.
        baselineCorrectionOn: Whether or not you want to correct for the baseline NL values
        gc_elution_on: Set to True if you need to integrate over GC curve, and account for change in ion NL over time
        gc_elution_times: time frames that each peak elutes at, to cull for
        cullAmount: A number of standard deviations from the mean. If an individual scan has the cullOn variable outside of this range, culls the scan; 
                    i.e. if cullOn is 'TIC*IT' and cullAmount is 3, culls scans where TIC*IT is more than 3 standard deviations from its mean. 
        NL_over_TIC: specific NL/TIC that designates what a "peak" should look like. default 0.1, currently not implemented
        isotopeList: A list of isotopes corresponding to the peaks extracted by FTStat, in the order they were extracted. This must be the same for each fragment. 
                    This is used to determine all ratios of interest, i.e. 13C/UnSub, and label them in the proper order. 

    Outputs: 
        A list of combined dataframes; in the 119/109 example above, it will output a list of two dataframes, [119, 109]
    where each dataframe combines the information for each substituted peak. This allows one to cull on the basis of different
    inputs (i.e. TIC*IT) as well as compute ratios (both of which are done by this function and _calcAppendRatios, which this 
    function calls). 
    '''
    DFList = []
    numberPeaksPerFragment = len(isotopeList)
    thisGCElutionTimeRange = []

    for peakIndex in range(len(peakDF)):
        if peakIndex % numberPeaksPerFragment == 0:
            df1 = peakDF[peakIndex].copy()
            
            #Set up a column to track total NL of peaks of fragment of interest for GC elution
            #and set up parameters for this specific fragment elution
            if gc_elution_on == True: #TODO: fix boolean processing
                thisGCElutionTimeRange = gc_elution_times[int(peakIndex / numberPeaksPerFragment)]

            #remove background
            if baselineCorrectionOn ==True:
                df1 = remove_background_NL(df1, thisGCElutionTimeRange)

            # calculate counts and add to the dataframe
            df1 = calculate_Counts_And_ShotNoise(df1)
           
            #Rename columns to keep track of them
            sub = isotopeList[0]
            df1.rename(columns={'mass':'mass'+sub,'counts':'counts'+sub,'absIntensity':'absIntensity'+sub,
                                'peakNoise':'peakNoise'+sub},inplace=True)
            df1['sumAbsIntensity'] = df1['absIntensity'+sub]
                       
            #helper variable to assign labels to final dataframe
            isotopeListIndex = 0

            #add additional dataframes
            for additionalDfIndex in range(numberPeaksPerFragment-1):
                df2 = peakDF[peakIndex + additionalDfIndex+1].copy()
                
                if baselineCorrectionOn ==True:
                    df2 = remove_background_NL(df2, thisGCElutionTimeRange)
                
                # calculate counts and add to the dataframe
                df2 = calculate_Counts_And_ShotNoise(df2)
           
                sub = isotopeList[additionalDfIndex+1]
                df2.rename(columns={'mass':'mass'+sub,'counts':'counts'+sub,'absIntensity':'absIntensity'+sub,
                                'peakNoise':'peakNoise'+sub},inplace=True)
                
                if gc_elution_on == True:
                    df1['sumAbsIntensity'] = df1['sumAbsIntensity'] + df2['absIntensity'+sub]
                
                #Drop duplicate information
                df2.drop(['retTime','tic','integTime','TIC*IT','ftRes','peakRes','peakBase'],axis=1,inplace=True) 
                          
                # merge 13C and 15N dataframes
                df1 = pd.merge_ordered(df1, df2,on='scanNumber',suffixes =(False,False))
                
                isotopeListIndex += 1

            #Checks each peak for values which were not recorded (e.g. due to low intensity) and fills in zeroes
            #I think this accomplishes the same thing as the zeroFilling in FTStat
            for string in isotopeList:
                df1.loc[df1['mass' + string].isnull(), 'mass' + string] = 0
                df1.loc[df1['absIntensity' + string].isnull(), 'absIntensity' + string] = 0
                df1.loc[df1['peakNoise' + string].isnull(), 'peakNoise' + string] = 0
                df1.loc[df1['counts' + string].isnull(), 'counts' + string] = 0 

            massStr = str(df1['mass'+isotopeList[0]].tolist()[0])

            #Cull zero scans
            if cullZeroScansOn == True:
                df1 = cull_Zero_Scans(df1)

            #Cull based on time frame for GC peaks
            if gc_elution_on == True and gc_elution_times != 0:
                df1= cull_On_GC_Peaks(df1, thisGCElutionTimeRange)

            #Cull reservoir measurments based on percent of maxNL of unsub peak
            if minNL_over_maxNL != 0:
                df1 = cull_On_Reservoir_Measurement(df1,minNL_over_maxNL)

            #Calculates ratio values and adds them to the dataframe. Weighted averages will be calculated in the next step
            df1 = calc_Append_Ratios(df1, isotopeList = isotopeList)
            #Given a key in the dataframe, culls scans outside specified multiple of standard deviation from the mean
            if cullOn != None:
                if cullOn not in list(df1):
                    raise Exception('Invalid Cull Input')
                maxAllowed = df1[cullOn].mean() + cullAmount * df1[cullOn].std()
                minAllowed = df1[cullOn].mean() - cullAmount * df1[cullOn].std()
                
                df1 = df1.drop(df1[(df1[cullOn] < minAllowed) | (df1[cullOn] > maxAllowed)].index)
            
            #Adds the combined dataframe to the output list
            DFList.append(df1)

        else:
            pass
    return DFList

def cull_Zero_Scans(df):
    '''
    Inputs:
        df: input dataframe to cull
    Outputs:
        culled df without zero
    '''
    df = df[~(df == 0).any(axis=1)]
    return df

def remove_background_NL(peakDF, gcElutionTime):
    '''
    Inputs:
        peakDF: data frame for this peak
        gcElutionTime: time frame that peak elutes at, which is necessary to distinguish the elution v baseline
    Outputs:
        culled df with background NL subtracted
    '''
    peakStartIndex = peakDF[peakDF['retTime'].between(gcElutionTime[0], gcElutionTime[1], inclusive=True)].index[0]
    baselineRows = peakDF[int(peakStartIndex - 10) : int(peakStartIndex - 5)]
    averageBaselineNL = baselineRows['absIntensity'].mean()
    peakDF['absIntensity'] = peakDF['absIntensity'] - averageBaselineNL
    
    return peakDF

def cull_On_GC_Peaks(df, gcElutionTimeFrame = (0,0)):
    '''
    Inputs: 
        df: input dataframe to cull
        gcElutionTimeFrame: elution of gc peaks, currently specified by the user
        NL_over_TIC: specific NL/TIC that designates what a "peak" should look like. default 0.1, currently not implemented
    Outputs: 
       culled df based on input elution times for the peaks
    '''
    # get the scan numbers for the retention  time frame
    if gcElutionTimeFrame != (0,0):

        #cull based on passed in retention time... 
        #this  could be a place in the future to cull based on automatic peak detection
        #Could also cull based on % threshhold of total peak height
        df = df[df['retTime'].between(gcElutionTimeFrame[0], gcElutionTimeFrame[1], inclusive=True)]
    return df

def cull_On_Reservoir_Measurement(df,minNL_over_maxNL=0.1):
    '''
    Implemented by Guannan Dong
    Inputs: 
        df: input dataframe to cull
        minNL_over_maxNL: reservoir measurements have a systematic varying NL. 
            (a rapid increase from the background level to max NL then a gradual decrease)
            minNL/maxNL designate a lower relative threshold for NL. Default 0.1.                
    Outputs: 
       culled df based on the acceptable minimum NL relative to a maximum NL.
    '''
    if minNL_over_maxNL != 0:
        maxNL = df['absIntensityUnSub'].max()
        maxNLIndex = df['absIntensityUnSub'].idxmax()               
        #use a boolean series to find the cutoff index for the acceptable min NL post NL max
        #this is to achieve a 'vertical cut' instead of a 'horizontal cut'
        s_postNLmax = df.loc[maxNLIndex:df.index[-1],'absIntensityUnSub']<minNL_over_maxNL*maxNL
        #cull any scan with NL < threshold, or exceeds the cutoff index
        try:
            df = df.drop(df[(df['absIntensityUnSub'] < minNL_over_maxNL*maxNL) | (df.index > s_postNLmax[s_postNLmax].index[0])].index)
        except:
            df = df[df['absIntensityUnSub'] > minNL_over_maxNL*maxNL]
            print("NL values post maximum are not culled")
    return df


def integrateTimeSeries(x, y, windowLength = 5, nanReplacer = 0.000001, slopeThreshhold =  0.08):
    '''
    Integrates a gaussian peak to return an area

    Inputs: 
        x = x values
        y = y values 
        windowLength = defines window used to calculate rolling slope
        nanReplacer = if slopes are NAN, what the values are replaced with in table. should be below slopeThreshhold
        slopeThreshhold = defines slope above which peak is defined
    Outputs: 
       area = area under integrated gaussian peak
    '''    
    data = {'x':x, 'y':y}

    #Integrate the data under a curve using two methods: (1) the trapezoid rule, and (2) sum of counts
    y_trap_int = scipy.integrate.trapz(y, x)
    y_sum_int = sum(y)

    return y_trap_int, y_sum_int
    
def calc_Raw_File_Output(dfList, isotopeList = ['13C','15N','UnSub'],omitRatios = []):
    '''
    For each ratio of interest, calculates mean, stdev, SErr, RSE, and ShotNoise based on counts. 
    Outputs these in a dictionary which organizes by fragment (i.e different entries for fragments at 119 and 109).
    
    Inputs:
        dfList: A list of merged data frames from the _combineSubstituted function. Each dataframe constitutes one fragment.
        isotopeList: A list of isotopes corresponding to the peaks extracted by FTStat, in the order they were extracted. 
                    This must be the same for each fragment. This is used to determine all ratios of interest, i.e. 13C/UnSub, and label them in the proper order. 
        omitRatios: A list of ratios to ignore. I.e. by default, the script will report 13C/15N ratios, which one may not care about. 
                    In this case, the list should be ['13C/15N','15N/13C'], including both versions, to avoid errors. 
         
    Outputs: 
        A dictionary giving mean, stdev, StandardError, relative standard error, and shot noise limit for all peaks.  
    '''
    #Initialize output dictionary 
    rtnDict = {}
      
    # iterate through each fragment
    for fragmentIndex in range(len(dfList)):
        
        #Adds the peak mass to the output dictionary
        key = round(dfList[fragmentIndex]['massUnSub'].mean())
        massStr = str(key)
        
        rtnDict[massStr] = {}
        
        for i in range(len(isotopeList)):
            for j in range(len(isotopeList)):
                if j>i:
                    if isotopeList[i] + '/' + isotopeList[j] in dfList[fragmentIndex]:
                        header = isotopeList[i] + '/' + isotopeList[j]
                    else:
                        try:
                            header = isotopeList[j] + '/' + isotopeList[i]
                        except:
                            raise Exception('Sorry, cannot find ratios for your input isotopeList')

                    if header in omitRatios:
                        print("Ratios omitted:" + header)
                        continue
                    else:
                            #This logic calculates ratios and statistics for each peak within each file, weighted by NL score for each scan
                            rtnDict[massStr][header] = {}
                            values = dfList[fragmentIndex][header]
                            weights = dfList[fragmentIndex]['absIntensityUnSub']
                            average = np.average(values, weights=weights)
                            rtnDict[massStr][header]['Ratio'] = average
                            rtnDict[massStr][header]['StDev'] = math.sqrt(
                                np.average((values-average)**2, weights=weights))
                            rtnDict[massStr][header]['StError'] = rtnDict[massStr][header]['StDev'] / \
                                np.power(len(dfList[fragmentIndex]), 0.5)
                            rtnDict[massStr][header]['RelStError'] = rtnDict[massStr][header]['StError'] / \
                                rtnDict[massStr][header]['Ratio']
                            
                            a = dfList[fragmentIndex]['counts' +
                                                      isotopeList[i]].sum()
                            b = dfList[fragmentIndex]['counts' +
                                                      isotopeList[j]].sum()
                            shotNoiseByQuad = np.power((1./a + 1./b), 0.5)
                            rtnDict[massStr][header]['ShotNoiseLimit by Quadrature'] = shotNoiseByQuad
                            averageTIC = np.mean(dfList[fragmentIndex]['tic'])
                            valuesTIC = dfList[fragmentIndex]['tic']
                            rtnDict[massStr][header]['TICVar'] = math.sqrt(
                                np.mean((valuesTIC-averageTIC)**2))/np.mean(valuesTIC)

                            averageTICIT = np.mean(
                                dfList[fragmentIndex]['TIC*IT'])
                            valuesTICIT = dfList[fragmentIndex]['TIC*IT']
                            rtnDict[massStr][header]['TIC*ITMean'] = averageTICIT
                            rtnDict[massStr][header]['TIC*ITVar'] = math.sqrt(
                                np.mean((valuesTICIT-averageTICIT)**2))/np.mean(valuesTICIT)

                            x = dfList[fragmentIndex]['retTime']
                            unsub_y = dfList[fragmentIndex]['counts' + isotopeList[i]]
                            sub_y = dfList[fragmentIndex]['counts' + isotopeList[j]]

                            #Integrate the curves based on the time frame chosen and return that R value
                            unsub_trap_integral, unsub_sum_integral = integrateTimeSeries(x, unsub_y, windowLength= WINDOW_LENGTH, nanReplacer= NAN_REPLACER, slopeThreshhold=SLOPE_THRESHHOLD)
                            sub_trap_integral, sub_sum_integral = integrateTimeSeries(x, sub_y, windowLength= WINDOW_LENGTH, nanReplacer= NAN_REPLACER, slopeThreshhold=SLOPE_THRESHHOLD)

                            #If trap rule boolean is on, use the trapezoid rule to integrate. OTherwise, sum the counts across the peak and use to calculate a ratio
                            if TRAP_RULE_BOOL == True:
                                sub_integral =sub_trap_integral 
                                unsub_integral = unsub_trap_integral
                            else:
                                sub_integral = sub_sum_integral 
                                unsub_integral = unsub_sum_integral
                            

                            try:
                                R_integrated = float(sub_integral) / float(unsub_integral)
                            except:
                                print("R_integrated gave NaN for peak:" + str(key))
                                R_integrated = 0

                            rtnDict[massStr][header]['Ratio_Integrated'] = R_integrated
    return rtnDict

def calc_Folder_Output(folderPath, cullOn=None, cullAmount=2,\
                       cullZeroScansOn=False, trapRuleOn = False, \
                       baselineSubstractionOn=False, gcElutionOn=False, \
                       gcElutionTimes = [], isotopeList = ['UnSub', '13C'], \
                       minNL_over_maxNL=0.1, omitRatios = [], fileCsvOutputPath=None):

    '''
    For each raw file in a folder, calculate mean, stdev, SErr, RSE, and ShotNoise based on counts. Outputs these in 
    a dictionary which organizes by fragment (i.e different entries for fragments at 119 and 109).  
    Inputs:
        folderPath: Path that all the .xslx raw files are in. Files must be in this format to be processed.
        cullOn: cull specific range of scans
        cullAmount: A number of standard deviations from the mean. If an individual scan has the cullOn variable outside of this range, culls the scan; 
                    i.e. if cullOn is 'TIC*IT' and cullAmount is 3, culls scans where TIC*IT is more than 3 standard deviations from its mean. 
        cullZeroScansOn: toggle to eliminate any scans with zero counts
        trapRuleOn: A toggle to specify whether to integrate by trapezoid rule (True) or by summing counts within a peak (False)
        baselineSubstractionOn: A toggle to specify whether or not to subtract baseline from NL values before calculating counts
        gcElutionOn: Specify whether you expect elution to change over time, so that you can calculate weighted averages
        gcElutionTimes: Time frames to cull the GC peaks for
        
        isotopeList: A list of isotopes corresponding to the peaks extracted by FTStat, in the order they were extracted. 
                    This must be the same for each fragment. This is used to determine all ratios of interest, i.e. 13C/UnSub, and label them in the proper order. 
        NL_over_TIC: Currently not used. Idea would be to cull any scans that the peak is below this percent of total TIC.
        omitRatios: A list of ratios to ignore. I.e. by default, the script will report 13C/15N ratios, which one may not care about. 
                    In this case, the list should be ['13C/15N','15N/13C'], including both versions, to avoid errors.
        fileCSVOutputPath: path name if you want to output each file as you process
        
    Outputs: 
        Output is a tuple:
        A dataframe giving mean, stdev, standardError, relative standard error, and shot noise limit for all peaks. 
        A dataframe with calculated statistics (average, stddev, stderror and rel std error)
        (Both the dataframes are also exported as csvs to the original input folder)
    '''

    ratio = "Ratio"
    stdev = "StdDev"
    rtnAllFilesDF = []
    header = ["FileNumber", "Fragment", "IsotopeRatio", "IntegratedIsotopeRatio", "Average", \
        "StdDev", "StdError", "RelStdError","TICVar","TIC*ITVar","TIC*ITMean", 'ShotNoise']
    #get all the file names in the folder with the same end 
    fileNames = [x for x in os.listdir(folderPath) if x.endswith(".xlsx")]
    peakNumber = 0

    #Process through each raw file added and calculate statistics for fragments of interest
    for i in range(len(fileNames)):
        thisFileName = str(folderPath + '/' + fileNames[i])
        print(thisFileName) #for debugging
        thesePeaks = import_Peaks_From_FTStatFile(thisFileName)
        thisPandas = convert_To_Pandas_DataFrame(thesePeaks)
        thisMergedDF = combine_Substituted_Peaks(peakDF=thisPandas,cullOn=cullOn, cullZeroScansOn = cullZeroScansOn, baselineCorrectionOn = baselineSubstractionOn, \
                gc_elution_on=gcElutionOn, gc_elution_times=gcElutionTimes, cullAmount=cullAmount, isotopeList=isotopeList, minNL_over_maxNL=minNL_over_maxNL)
        thisOutput = calc_Raw_File_Output(thisMergedDF, isotopeList, omitRatios)
        keys = list(thisOutput.keys())
        peakNumber = len(keys)

        for peak in range(peakNumber):
            #key is each peak within the dictionary
            isotopeRatios = list(thisOutput[keys[peak]].keys())
            #subkey is each isotope ratio info for each peak
            for isotopeRatio in range(len(isotopeRatios)):
                thisPeak = keys[peak]
                thisRatio = isotopeRatios[isotopeRatio]
                #add subkey to each separate df for isotope specific 
                thisRVal = thisOutput[thisPeak][thisRatio]["Ratio"]
                thisRIntegratedVal = thisOutput[thisPeak][thisRatio]["Ratio_Integrated"]
                thisStdDev = thisOutput[thisPeak][thisRatio]["StDev"]
                thisStError = thisOutput[thisPeak][thisRatio]["StError"] 
                thisRelStError = thisOutput[thisPeak][thisRatio]["RelStError"]
                thisTICVar = thisOutput[thisPeak][thisRatio]["TICVar"] 
                thisTICITVar = thisOutput[thisPeak][thisRatio]["TIC*ITVar"]
                thisTICITMean = thisOutput[thisPeak][thisRatio]["TIC*ITMean"]
                thisShotNoise = thisOutput[thisPeak][thisRatio]["ShotNoiseLimit by Quadrature"]
                thisRow = [thisFileName, thisPeak, thisRatio, thisRIntegratedVal, thisRVal, \
                    thisStdDev, thisStError,thisRelStError,thisTICVar,thisTICITVar,thisTICITMean, thisShotNoise] 
                rtnAllFilesDF.append(thisRow)

    rtnAllFilesDF = pd.DataFrame(rtnAllFilesDF)
    # set the header row as the df header
    rtnAllFilesDF.columns = header 

    #sort by fragment and isotope ratio, output to csv
    rtnAllFilesDF = rtnAllFilesDF.sort_values(by=['Fragment', 'IsotopeRatio'], axis=0, ascending=True)
    rtnAllFilesDF.to_csv(str(folderPath + '/' + "all_data_output.csv"), index = False, header=True)

    #we can now calculate average, stdev, relstdev for each fragment across replicate measurements 
    if len(fileNames)>1: #only calculate  stats if there is more than one file
        avgDF = rtnAllFilesDF.groupby(['Fragment', 'IsotopeRatio'])["Average"].mean()
        integratedAvgDF = rtnAllFilesDF.groupby(['Fragment', 'IsotopeRatio'])["IntegratedIsotopeRatio"].mean()
        countDF = rtnAllFilesDF.groupby(['Fragment', 'IsotopeRatio'])["Average"].count()
        stdDF = rtnAllFilesDF.groupby(['Fragment', 'IsotopeRatio'])["Average"].std()
        trapStdDF = rtnAllFilesDF.groupby(['Fragment', 'IsotopeRatio'])["IntegratedIsotopeRatio"].std()
        sqrtCountDF = np.power(countDF, 0.5)
        stdErrorDF = np.divide(stdDF, sqrtCountDF)
        trapStdErrorDF = np.divide(trapStdDF, sqrtCountDF)
        relStdErrorDF = np.divide(stdErrorDF, avgDF)
        trapRelStdErrorDF = np.divide(trapStdErrorDF, integratedAvgDF)

    statsDF = pd.DataFrame([avgDF, integratedAvgDF, countDF, stdDF, stdErrorDF, \
        relStdErrorDF, trapStdDF, trapStdErrorDF, trapRelStdErrorDF], \
            index=["ScanByScan Avg R Val", "Integrated Avg R val", "N", "ScanStdDev", "ScanStdError", \
                "ScanRelStdError", "IntStdDev", "IntStdError", "IntRelStdError"]) 
    
    #output results to csv
    statsDF.to_csv(str(folderPath + '/' + "stats_output.csv"), index = True, header=True)
   
    return rtnAllFilesDF, statsDF